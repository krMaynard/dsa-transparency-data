#!/usr/bin/env python3
"""Build microsoft-lerr.json from Microsoft's Law Enforcement Requests Report.

Microsoft publishes its **Law Enforcement Requests Report** (LERR) as one XLSX
workbook per half-year since 2013, linked from
https://www.microsoft.com/en-us/corporate-responsibility/law-enforcement-requests-report
at stable asset slugs (https://cdn-dynmedia-1.microsoft.com/is/content/microsoftcorp/
Microsoft-LERR-<YYYY>-<HN>). Each workbook reports, per country, the number of
legal demands received, the accounts/users specified, and the four disclosure
outcomes (content / only non-content subscriber-transactional data / no data
found / rejected).

The workbook layout changed over the years — three column geometries and an
evolving sheet split:

  * 2013: two sheets, ``MSFT`` (all Microsoft services incl. Skype) + ``Skype``
    (Skype separately); country in column A, counts from column B.
  * 2014–2016: a single combined sheet; country in column A, counts from C.
  * 2017: ``Total - Criminal`` + ``Total - Emergencies`` sheets (same geometry).
  * 2018–2021: ``Criminal`` / ``Emergencies`` / ``Civil Legal Requests``
    sheets, everything shifted one column right (country in B, counts from D).
  * 2022→: ``Civil`` / ``Criminal`` / ``Emergencies``, back to the 2014
    geometry.

The parser locates each metric's column from the descriptive header texts
(stable across all eras) rather than hardcoding per-file maps, reads Excel's
**cached computed values**
(``data_only=True`` — several eras store every figure as a formula referencing
other sheets), skips the percentage columns (derivable from the counts), and
**validates the extracted per-country sums against the workbook's own TOTAL
row**. The one exception: on civil sheets the workbook's own requests TOTAL
de-duplicates requests spanning multiple countries, so it is smaller than the
per-country column sum — that metric is excluded from validation there. Note
also that on civil sheets the four outcome columns count **accounts** (they
sum to the accounts-specified figure per row), not requests.

Output is one **tidy-long** table — one row per measured value:

  period, section, country, metric, unit, value

``section`` is the report split the row came from (``combined`` / ``skype`` /
``criminal`` / ``emergencies`` / ``civil``) — the split changes across eras
(``combined`` runs 2013–2016; ``criminal``/``emergencies`` from 2017; ``civil``
from 2018), so **pin a section before aggregating** and never sum sections with
``combined``. ``skype`` overlaps ``combined`` in 2013 (MSFT figures include
Skype). Metrics: ``requests`` / ``accounts_specified`` / ``disclosed_content``
/ ``disclosed_noncontent`` / ``no_data_found`` / ``rejected``, all ``count``.

Deterministic: builds purely from the archived raw/ workbooks (rows sorted); no
wall-clock. ``--download`` refreshes raw/ from the live asset URLs. Needs
``openpyxl``.
"""
from __future__ import annotations

import argparse
import json
import os
import re
import urllib.request

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
RAW_DIR = os.path.join(HERE, "raw")
OUT_JSON = os.path.join(HERE, "microsoft-lerr.json")

ASSET_URL = "https://cdn-dynmedia-1.microsoft.com/is/content/microsoftcorp/Microsoft-LERR-{year}-{half}"
FIRST_YEAR = 2013
LAST = (2025, "H1")  # newest published period; extend as new halves appear

COLUMNS = ["period", "section", "country", "metric", "unit", "value"]

SECTIONS = {
    "MSFT": "combined",
    "All Microsoft Services": "combined",
    "Combined LERR": "combined",
    "LERR": "combined",
    "Sheet1": "combined",
    "TOTAL": "combined",
    "Skype": "skype",
    "Total - Criminal": "criminal",
    "Criminal": "criminal",
    "Total - Emergencies": "emergencies",
    "Emergencies": "emergencies",
    # 2017-H2 names its criminal/emergencies sheets "<name> (2)"; the plain
    # "TOTAL" (2016-H2) is the combined report, but "TOTAL (2)" sits alongside
    # an Emergencies sheet, making it the criminal split (as in 2017-H1).
    "TOTAL (2)": "criminal",
    "Emergencies (2)": "emergencies",
    "Civil Legal Requests": "civil",
    "Civil": "civil",
}

# The 2017-H2 workbook's civil sheet is titled "Civil Legal Request
# Disclosures, January 2017 - June 2017" — it carries the H1 figures (civil
# reporting debuted with a catch-up sheet). Every other workbook's civil sheet
# matches its own period.
PERIOD_OVERRIDES = {("2017-H2", "civil"): "2017-H1"}

# Publisher inconsistencies: sheets whose hand-typed TOTAL cell disagrees with
# the sum of its own country rows. Keyed by (period, section, metric) →
# (our extracted sum, the workbook's stated total); the extracted per-country
# rows are kept and the pinned values keep the guard strict for re-runs.
KNOWN_TOTAL_MISMATCHES = {
    # The 2025-H1 Criminal TOTAL row carries hand-typed literals whose
    # content/non-content split disagrees with the sheet's own country rows by
    # a complementary ±2 (all other metrics reconcile exactly — two requests
    # were evidently re-categorised after the totals were typed).
    ("2025-H1", "criminal", "disclosed_content"): (1356, 1358),
    ("2025-H1", "criminal", "disclosed_noncontent"): (16945, 16943),
}

# Column-header patterns → metric. The column geometry shifts between eras
# (and one sheet is unique), but the descriptive headers are stable, so the
# value columns are located by header text: the requests/accounts figures sit
# in the labelled column itself; each outcome label heads a %/# pair, with the
# count one column to the right of the label.
HEADER_METRICS = [
    (re.compile(r"total number of\s+(law enforcement\s+)?requests", re.I), "requests", 0),
    (re.compile(r"accounts", re.I), "accounts_specified", 0),
    (re.compile(r"disclosure of content", re.I), "disclosed_content", 1),
    (re.compile(r"subscriber", re.I), "disclosed_noncontent", 1),
    (re.compile(r"no\s+data(\s+found)?\s*\)?$", re.I), "no_data_found", 1),
    (re.compile(r"reject", re.I), "rejected", 1),
]


def periods() -> list[tuple[int, str]]:
    out = []
    for year in range(FIRST_YEAR, LAST[0] + 1):
        for half in ("H1", "H2"):
            if (year, half) <= LAST:
                out.append((year, half))
    return out


def download() -> None:
    os.makedirs(RAW_DIR, exist_ok=True)
    for year, half in periods():
        url = ASSET_URL.format(year=year, half=half)
        req = urllib.request.Request(url, headers={"User-Agent": "dsa-transparency-data/1.0"})
        with urllib.request.urlopen(req, timeout=120) as resp:
            blob = resp.read()
        path = os.path.join(RAW_DIR, f"LERR-{year}-{half}.xlsx")
        with open(path, "wb") as f:
            f.write(blob)
        print(f"downloaded {url} ({len(blob)} bytes)")


def _clean_country(cell: object) -> str:
    name = str(cell).strip()
    name = re.sub(r"[*†‡]+$", "", name).strip()  # strip footnote markers
    return name


def _count(cell: object) -> int | None:
    """A literal numeric cell -> int; formulas/blank/text -> None."""
    if isinstance(cell, bool) or cell is None:
        return None
    if isinstance(cell, (int, float)):
        return int(round(cell))
    s = str(cell).strip().replace(",", "")
    if s.startswith("=") or s == "":
        return None
    try:
        return int(round(float(s)))
    except ValueError:
        return None


def parse_sheet(ws, period: str, section: str) -> list[tuple[str, str, str, str, str, int]]:
    grid = [list(r) for r in ws.iter_rows(values_only=True)]

    # Locate the TOTAL header row; the column it sits in is the country column.
    total_rc = None
    for ri, row in enumerate(grid[:12]):
        for ci, cell in enumerate(row[:3]):
            if isinstance(cell, str) and cell.strip() == "TOTAL":
                total_rc = (ri, ci)
                break
        if total_rc:
            break
    if not total_rc:
        raise ValueError(f"{period} {section}: no TOTAL row found")
    total_ri, country_col = total_rc

    # Locate each metric's value column from the descriptive headers above the
    # TOTAL row (geometry shifts between eras; the header texts don't).
    cols: dict[str, int] = {}
    for row in grid[:total_ri]:
        for ci, cell in enumerate(row):
            if not isinstance(cell, str):
                continue
            text = " ".join(cell.split())
            for pat, metric, delta in HEADER_METRICS:
                if metric not in cols and pat.search(text):
                    cols[metric] = ci + delta
    missing = [m for _, m, _ in HEADER_METRICS if m not in cols]
    if missing:
        raise ValueError(f"{period} {section}: header columns not found for {missing}")
    offsets = [(m, col - country_col) for m, col in cols.items()]

    rows: list[tuple[str, str, str, str, str, int]] = []
    sums: dict[str, int] = {m: 0 for m, _ in offsets}
    blanks = 0
    for row in grid[total_ri + 1:]:
        cell = row[country_col] if country_col < len(row) else None
        if cell is None or str(cell).strip() == "":
            blanks += 1
            if blanks >= 2:
                break
            continue
        blanks = 0
        country = _clean_country(cell)
        if country.upper() == "TOTAL" or len(country) < 2:
            continue
        got_any = False
        for metric, off in offsets:
            v = _count(row[country_col + off]) if 0 <= country_col + off < len(row) else None
            if v is not None:
                rows.append((period, section, country, metric, "count", v))
                sums[metric] += v
                got_any = True
        if not got_any:  # a footnote line, not a data row
            continue

    # Validate against the workbook's own (cached) TOTAL row. Civil sheets'
    # requests TOTAL de-duplicates multi-country requests, so it legitimately
    # undercounts the per-country column sum — skip that one check.
    for metric, off in offsets:
        if section == "civil" and metric == "requests":
            continue
        total_row = grid[total_ri]
        stated = (_count(total_row[country_col + off])
                  if 0 <= country_col + off < len(total_row) else None)
        if stated is not None and stated != sums[metric]:
            known = KNOWN_TOTAL_MISMATCHES.get((period, section, metric))
            if known == (sums[metric], stated):
                continue
            raise ValueError(
                f"{period} {section} {metric}: extracted sum {sums[metric]} != stated TOTAL {stated}")

    if not rows:
        raise ValueError(f"{period} {section}: no data rows extracted")
    return rows


def build() -> list[tuple[str, str, str, str, str, int]]:
    rows: list[tuple[str, str, str, str, str, int]] = []
    for year, half in periods():
        path = os.path.join(RAW_DIR, f"LERR-{year}-{half}.xlsx")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            period = f"{year}-{half}"
            for name in wb.sheetnames:
                if name not in SECTIONS:
                    raise ValueError(f"{period}: unknown sheet {name!r} — extend SECTIONS")
                if name == "Sheet1" and len(wb.sheetnames) > 1:
                    # "Sheet1" is the combined report only when it is the sole
                    # sheet (2016-H1); 2024-H2 also carries a tiny scratch Sheet1
                    # (a 2x4 chart-percentage pivot) next to the real sheets.
                    continue
                section = SECTIONS[name]
                sheet_period = PERIOD_OVERRIDES.get((period, section), period)
                rows.extend(parse_sheet(wb[name], sheet_period, section))
        finally:
            wb.close()
    rows.sort()
    return rows


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--download", action="store_true",
                    help="refresh raw/ from the live report assets before building")
    args = ap.parse_args()

    if args.download:
        download()

    rows = build()
    pds = sorted({r[0] for r in rows})
    out = {
        "source": "https://www.microsoft.com/en-us/corporate-responsibility/law-enforcement-requests-report",
        "coverage": f"{pds[0]}..{pds[-1]}",
        "columns": COLUMNS,
        "rows": [list(r) for r in rows],
    }
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, separators=(",", ":"))
        f.write("\n")
    sections = sorted({r[1] for r in rows})
    print(f"wrote {OUT_JSON}: {len(rows)} rows, {len(pds)} periods "
          f"({pds[0]}..{pds[-1]}), sections: {', '.join(sections)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
