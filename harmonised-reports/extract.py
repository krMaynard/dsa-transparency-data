#!/usr/bin/env python3
"""Extract DSA harmonised-template transparency reports into a normalised form.

The EU harmonised machine-readable template (Implementing Regulation (EU)
2024/2835) is an 11-section workbook — the same tables 1-11 used by the VLOP
dataset. Platforms ship it as a single multi-sheet .xlsx/.xls or as a zip of one
CSV per section; sheet names are sometimes localised (DE/FR), but the section
*order* (1..11) is fixed, so we map by position.

For each source under raw/, this writes raw/../extracted/<platform>/NN_<section>.csv
(one canonical CSV per section) and a manifest.json summarising what was found.

Re-run after adding files to raw/:  python3 extract.py
"""
from __future__ import annotations

import csv
import datetime
import io
import json
import os
import zipfile

HERE = os.path.dirname(os.path.abspath(__file__))
RAW = os.path.join(HERE, "raw")
OUT = os.path.join(HERE, "extracted")

# Canonical section names, in the fixed template order (1..11).
SECTIONS = [
    "report_identification", "categories_names", "member_states_orders",
    "notices", "own_initiative_illegal", "own_initiative_TC",
    "appeals_and_recidivism", "automated_means", "human_resources",
    "AMAR", "qualitative",
]

# platform slug -> (source filename in raw/, kind). kind: xlsx | xls | zipcsv
SOURCES = {
    "aboutyou":  ("aboutyou.csv.xlsx", "xlsx"),
    "aboutyou2": ("aboutyou2.xlsx", "xlsx"),  # AboutYou follow-on period (Dec 2025)
    "dailymotion": ("dailymotion.xlsx", "xlsx"),
    "carrefour": ("carrefour.xlsx", "xlsx"),
    "linkedin":  ("linkedin.zip", "zipcsv"),
    "manomano":  ("manomano.xlsx", "xlsx"),
    "pinterest": ("pinterest.zip", "zipcsv"),
    "veepee":    ("veepee.xlsx", "xlsx"),
    "vinted":    ("vinted.xlsx", "xlsx"),
    "webde":     ("webde.xlsx", "xlsx"),
    "wikipedia": ("wikipedia.xls", "xls"),
    # Added from the hub landing-page sweep (discover_hubs.py / download_hubs.py).
    "ceneo":       ("ceneo.xlsx", "xlsx"),
    "cloudflare":  ("cloudflare.xlsx", "xlsx"),
    "duckduckgo":  ("duckduckgo.xlsx", "xlsx"),
    "expedia":     ("expedia.xlsx", "xlsx"),
    "hometogo":    ("hometogo.xlsx", "xlsx"),
    "hostelworld": ("hostelworld.xlsx", "xlsx"),
    "hostinger":   ("hostinger.xlsx", "xlsx"),
    "hotelscom":   ("hotelscom.xlsx", "xlsx"),
    "imdb":        ("imdb.zip", "zipcsv"),
    "konami":      ("konami.xlsx", "xlsx"),
    "lilo":        ("lilo.xlsx", "xlsx"),
    "matchgroup":  ("matchgroup.xlsx", "xlsx"),
    "niantic":     ("niantic.xlsx", "xlsx"),
    "qwant":       ("qwant.xlsx", "xlsx"),
    "roblox":      ("roblox.xlsx", "xlsx"),
    "shopify":     ("shopify.xlsx", "xlsx"),
    "skroutz":     ("skroutz.xlsx", "xlsx"),
    "vrbo":        ("vrbo.xlsx", "xlsx"),
    "yahoo":       ("yahoo.zip", "zipcsv"),
    # Added from the Zendesk help-center API sweep (browser-blocked landing pages
    # whose attachment files resolve via /api/v2/help_center/.../attachments.json).
    "bumble":      ("bumble.xlsx", "xlsx"),
    "grindr":      ("grindr.xlsx", "xlsx"),
    "vestiaire":   ("vestiaire.xlsx", "xlsx"),
    "whatnot":     ("whatnot.xlsx", "xlsx"),
    "depop":       ("depop.xlsx", "xlsx"),
    "nexon":       ("nexon.xlsx", "xlsx"),
    "nintendo":    ("nintendo.xlsx", "xlsx"),
    "squareenix":  ("squareenix.xlsx", "xlsx"),
    "alibabacloud": ("alibabacloud.zip", "zipcsv"),
    # One zip holding a full 11-section report per game -> one extracted dir each.
    "miniclip":    ("miniclip.zip", "zipmulti"),
    # Format-variant reports: standard template *content*, but the sheet/file
    # names don't carry the canonical section number (LINE) or carry a
    # *renumbered* one (Discord). Mapped by name via SHEET_MAP (see below).
    "line":        ("line.xlsx", "xlsx"),
    "discord":     ("discord.zip", "zipcsv"),
    "gemini":      ("gemini.zip", "zipcsv"),
    "notebooklm":  ("notebooklm.zip", "zipcsv"),
    # Batch of Google services (H2 2025, non-VLOP, canonical numbered-CSV zips).
    "pubdev":                    ("pubdev.zip", "zipcsv"),
    "workspace":                 ("workspace.zip", "zipcsv"),
    "wallet-api":                ("wallet-api.zip", "zipcsv"),
    "looker":                    ("looker.zip", "zipcsv"),
    "google-pay-api":            ("google-pay-api.zip", "zipcsv"),
    "google-ai-developer-forum": ("google-ai-developer-forum.zip", "zipcsv"),
    "gdp-forums":                ("gdp-forums.zip", "zipcsv"),
    "vacation-rentals":          ("vacation-rentals.zip", "zipcsv"),
    "google-hotels":             ("google-hotels.zip", "zipcsv"),
    "google-flights":            ("google-flights.zip", "zipcsv"),
    "tenor":                     ("tenor.zip", "zipcsv"),
    "google-photos":             ("google-photos.zip", "zipcsv"),
    "manufacturer-center":       ("manufacturer-center.zip", "zipcsv"),
    "google-help-support":       ("google-help-support.zip", "zipcsv"),
    "google-news":               ("google-news.zip", "zipcsv"),
    "google-public-dns":         ("google-public-dns.zip", "zipcsv"),
    "waze":                      ("waze.zip", "zipcsv"),
    "google-ads":                ("google-ads.zip", "zipcsv"),
    "google-classroom":          ("google-classroom.zip", "zipcsv"),
    "chrome-web-store":          ("chrome-web-store.zip", "zipcsv"),
    "google-cloud-storage":      ("google-cloud-storage.zip", "zipcsv"),
    "colab":                     ("colab.zip", "zipcsv"),
    "fitbit":                    ("fitbit.zip", "zipcsv"),
    # Reddit — non-VLOP DSA report (2025), canonical numbered-CSV zip.
    "reddit":                    ("reddit.zip", "zipcsv"),
    "just-eat-takeaway":         ("just-eat-takeaway.zip", "zipcsv"),
    # Browser-fetched from the SOURCES-NEEDING-BROWSER §C backlog. Each is the
    # canonical numbered 11-section Annex I template (jeuxvideo in French, but the
    # sheets carry the 1..11 numbering the extractor keys on).
    "riot":        ("riot.xlsx", "xlsx"),        # Riot Games — 2025
    "flickr":      ("flickr.xlsx", "xlsx"),      # Flickr (SmugMug) — 2025 Annex I
    "jeuxvideo":   ("jeuxvideo.xlsx", "xlsx"),   # Jeuxvideo.com (Webedia) — 2025
    "glassdoor":   ("glassdoor.xlsx", "xlsx"),   # Glassdoor — CY2024
    "akamai":      ("akamai-2025-h2.xlsx", "xlsx"),  # Akamai — H2 2025
    "upwork":      ("upwork-2025.xlsx", "xlsx"),     # Upwork — CY2025
    "xkom":        ("xkom.xlsx", "xlsx"),            # x-kom — CY2024
    "vimeo-2024":  ("vimeo-2024.xlsx", "xlsx"),      # Vimeo — CY2024
    "vimeo-2025":  ("vimeo-2025.xlsx", "xlsx"),      # Vimeo — CY2025
}

# Sources whose sheet/file names can't be mapped by a parsed section number,
# because they're unnumbered or renumbered. Map each sheet/file to a canonical
# section (1..11) by a lower-cased name substring instead. List the most
# specific keys first; the first match wins. Sections not listed stay empty.
SHEET_MAP = {
    # LINE (LY Corp) condenses the template into 5 named, unnumbered sheets.
    # "own_initiative" carries the illegal-content category x restriction-type
    # grid with no surface column -> section 5 (own-initiative, illegal content);
    # "statements" is the free-text indicator/value table -> section 11.
    "line": [
        ("report_identification", 1), ("member_states_orders", 3),
        ("notices", 4), ("own_initiative", 5), ("statements", 11),
    ],
    # Discord omits own-initiative-illegal (5), human resources (9) and AMAR (10),
    # then renumbers what remains 5..8 — so its "5. Own Initiative TC" is really
    # section 6, "6. Appeals" is 7, "7. Automated Means" is 8, "8. Qualitative" is
    # 11. Map by name so the renumbering doesn't land rows in the wrong table.
    "discord": [
        ("report identification", 1), ("categories names", 2),
        ("member states orders", 3), ("notices", 4),
        ("own initiative tc", 6), ("appeals and recidivism", 7),
        ("automated means", 8), ("qualitative", 11),
    ],
    # Upwork is a non-VLOP and omits the VLOP-only human-resources and AMAR
    # sheets. Its remaining nine sheets are named but unnumbered, so map them
    # explicitly rather than shifting the final qualitative sheet into table 9.
    "upwork": [
        ("report_identification", 1), ("categories_names", 2),
        ("member_states_orders", 3), ("notices", 4),
        ("own_initiative_illegal", 5), ("own_initiative_tc", 6),
        ("appeals_and_recidivism", 7), ("automated_means", 8),
        ("qualitative", 11),
    ],
}

import re as _re
# Leading/embedded section number, e.g. "1_report…", "… - 10. AMAR", "Part 11".
_SEC_NUM = _re.compile(r"(?:^|[-\s_/])(\d{1,2})[._\s]")


def _section_index(name: str) -> int | None:
    nums = [int(n) for n in _SEC_NUM.findall(name) if 1 <= int(n) <= 11]
    return nums[-1] if nums else None


def _to_canonical(named: list[tuple[str, list[list[str]]]]) -> list[list[list[str]]]:
    """Map (sheet/file name, rows) pairs to the 11 canonical section slots.

    Prefer the section number parsed from the name (robust to omitted sections,
    e.g. a non-VLOP that ships 1-8 + 11). Fall back to positional order only when
    no name carries a number and exactly 11 are present.
    """
    idxs = [_section_index(n) for n, _ in named]
    if all(i is not None for i in idxs) and len(set(idxs)) == len(idxs):
        slots: list[list[list[str]]] = [[] for _ in SECTIONS]
        for i, (_, rows) in zip(idxs, named):
            if i and 1 <= i <= 11:
                slots[i - 1] = rows
        return slots
    if len(named) == len(SECTIONS):  # positional fallback (unnumbered, full set)
        return [rows for _, rows in named]
    # Best effort: place the numbered ones, leave the rest empty.
    slots = [[] for _ in SECTIONS]
    for i, (_, rows) in zip(idxs, named):
        if i and 1 <= i <= 11:
            slots[i - 1] = rows
    return slots


def _to_canonical_mapped(named: list[tuple[str, list[list[str]]]],
                         mapping: list[tuple[str, int]]) -> list[list[list[str]]]:
    """Map (sheet/file name, rows) pairs to the 11 canonical slots by matching a
    lower-cased name substring against `mapping` (first match wins), for sources
    whose names don't carry a usable canonical section number."""
    slots: list[list[list[str]]] = [[] for _ in SECTIONS]
    for name, rows in named:
        low = name.lower()
        for key, idx in mapping:
            if key in low and 1 <= idx <= 11:
                slots[idx - 1] = rows
                break
    return slots


def _clean(v: object) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    # Excel stores a "median time" duration as a serial; openpyxl (data_only)
    # surfaces it as a datetime in 1900 (e.g. 1900-01-03 22:52:00). Render it as a
    # real duration measured from the Excel epoch, not a bogus 1900 calendar date.
    if isinstance(v, datetime.datetime) and v.year <= 1900:
        return str(v - datetime.datetime(1899, 12, 30))
    if isinstance(v, datetime.time):
        return v.strftime("%H:%M:%S")
    return str(v).strip()


def read_xlsx(path: str) -> list[tuple[str, list[list[str]]]]:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = [[_clean(c) for c in row] for row in ws.iter_rows(values_only=True)]
        while rows and not any(rows[-1]):  # trim trailing empty rows
            rows.pop()
        sheets.append((name, rows))
    wb.close()
    return sheets


def read_xls(path: str) -> list[tuple[str, list[list[str]]]]:
    import xlrd
    from xlrd.xldate import xldate_as_datetime
    wb = xlrd.open_workbook(path)
    sheets = []
    for sh in wb.sheets():
        rows = []
        for r in range(sh.nrows):
            row = []
            for c in range(sh.ncols):
                # Convert legacy-xls date serials to ISO so they match the xlsx
                # files' "YYYY-MM-DD HH:MM:SS" rather than leaking 45839-style serials.
                if sh.cell_type(r, c) == xlrd.XL_CELL_DATE:
                    dt = xldate_as_datetime(sh.cell_value(r, c), wb.datemode)
                    row.append(dt.strftime("%Y-%m-%d %H:%M:%S"))
                else:
                    row.append(_clean(sh.cell_value(r, c)))
            rows.append(row)
        while rows and not any(rows[-1]):
            rows.pop()
        sheets.append((sh.name, rows))
    return sheets


_ADS_RE = _re.compile(r"\d+_.*_ads\.csv$", _re.I)
# Only sections 6/7/8 carry a surface dimension downstream (t6/t7/t8 in the API
# star schema). An ads-surface sub-breakdown is only meaningful for those; an
# _Ads file for any other section can't be held as a surface and must not be
# folded (it would concatenate Core+Ads rows into a surface-less table and
# silently double-count). Keep this set in sync with the API seeder.
_SURFACE_SECTIONS = {6, 7, 8}


def _stamp_surface(rows: list[list[str]], surface: str) -> list[list[str]]:
    """Append a trailing 'Surface' column ('Surface' header + `surface` on each
    data row), so a downstream loader can read the surface from the last cell."""
    if not rows:
        return rows
    return [rows[0] + ["Surface"]] + [r + [surface] for r in rows[1:]]


def _merge_ads_surfaces(
        pairs: list[tuple[str, list[list[str]]]]) -> list[tuple[str, list[list[str]]]]:
    """Some Google reports (Hotels, Workspace) split sections 6-8 into a base file
    plus a sibling 'NN_<section>_Ads.csv'. Both label their leading 'Applicability'
    column 'All', but in Google's own VLOP reporting these are the additive,
    non-overlapping **Core** and **Ads** surfaces — there is no aggregate 'All'
    surface row (summing one would double-count the other). So fold each _Ads file
    into its base section, stamping a trailing 'Surface' column: the base rows are
    'Core', the _Ads rows 'Ads'. Reports without an _Ads sibling are returned
    unchanged (their single 'All' surface is the whole service)."""
    ads_by_sec: dict[int, tuple[str, list[list[str]]]] = {}
    base: list[tuple[str, list[list[str]]]] = []
    for name, rows in pairs:
        if not _ADS_RE.match(name):
            base.append((name, rows))
            continue
        si = _section_index(name)
        if si in _SURFACE_SECTIONS:
            if si in ads_by_sec:
                # Two ads files map to the same section — keeping only the last
                # would silently drop the first, so warn rather than lose data.
                print(f"WARNING: multiple ads-surface files for section {si} "
                      f"({ads_by_sec[si][0]!r} and {name!r}); keeping {name!r}")
            ads_by_sec[si] = (name, rows)
        else:
            # An ads-surface file for a section that has no surface dimension
            # (or whose number didn't parse). Folding it would double-count into
            # a surface-less table, so drop it and warn loudly — a human should
            # extend extract.py + the schema rather than ingest it blindly. (It is
            # neither folded nor kept, so it can't collide in _to_canonical.)
            where = f"section {si}" if si is not None else "an unparseable section"
            print(f"WARNING: ignoring ads-surface file {name!r} for {where} "
                  f"— only sections {sorted(_SURFACE_SECTIONS)} carry a surface dimension")
    # Rebuild from `base` (not `pairs`): this both folds the matched ads sections
    # and excludes any skipped/dropped ads file. With no ads files at all, `base`
    # is every input pair in order, so the output equals the input.
    out: list[tuple[str, list[list[str]]]] = []
    consumed: set[int] = set()
    for name, rows in base:
        si = _section_index(name)
        ads_entry = ads_by_sec.get(si) if si is not None else None
        if ads_entry is None:
            out.append((name, rows))
            continue
        _, ads_rows = ads_entry
        merged = _stamp_surface(rows, "Core")
        merged += _stamp_surface(ads_rows, "Ads")[1:]  # ads data rows (drop header)
        out.append((name, merged))
        consumed.add(si)
    # Defensive: an _Ads file with no base sibling (not present in our data) —
    # keep it as its own Ads-surface section (under its original name) rather
    # than silently dropping rows.
    for si, (ads_name, rows) in ads_by_sec.items():
        if si not in consumed:
            out.append((ads_name, _stamp_surface(rows, "Ads")))
    return out


def read_zipcsv(path: str) -> list[tuple[str, list[list[str]]]]:
    """Read the per-section CSVs from a zip as (filename, rows) pairs."""
    out = []
    with zipfile.ZipFile(path) as z:
        for info in z.namelist():
            base = os.path.basename(info)
            # Skip non-CSVs and macOS metadata (the __MACOSX/ dir and the
            # AppleDouble "._name" sidecar files, which are binary, not CSV).
            if (not base.lower().endswith(".csv") or info.startswith("__MACOSX")
                    or base.startswith("._")):
                continue
            with z.open(info) as f:
                data = f.read()
            # Most zips are UTF-8; some (e.g. Alibaba Cloud) are CP1252 — decode
            # strict UTF-8 first, fall back to CP1252 so smart quotes/ellipses
            # survive instead of becoming U+FFFD replacement chars.
            try:
                text = data.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = data.decode("cp1252")
            # Parse via StringIO (not splitlines) so newlines *inside* quoted
            # fields — common in the qualitative section — don't split a row.
            rows = list(csv.reader(io.StringIO(text)))
            while rows and not any(c.strip() for c in rows[-1]):
                rows.pop()
            out.append((base, rows))
    # Fold any ads-surface sub-breakdown files (Hotels, Workspace) into their base
    # section, tagged Core/Ads, instead of dropping them.
    return _merge_ads_surfaces(out)


READERS = {"xlsx": read_xlsx, "xls": read_xls, "zipcsv": read_zipcsv}


def _trim_cols(rows: list[list[str]]) -> list[list[str]]:
    """Drop trailing all-empty columns (spreadsheets pad rows to the sheet width)."""
    width = 0
    for r in rows:
        last = max((i + 1 for i, c in enumerate(r) if c), default=0)
        width = max(width, last)
    return [r[:width] for r in rows]


# Indicator labels in section 1, across the locales we ingest (EN/DE/FR/EL).
_PROVIDER = ("service provider", "diensteanbieters", "fournisseur", "παρόχου",
             "dostawcy usług")
_START = ("starting date", "beginn des berichtszeitraums", "date de début",
          "début", "έναρξης", "rozpoczęcia okresu sprawozdawczego")
_END = ("ending date", "ende des berichtszeitraums", "date de fin", "fin", "λήξης",
        "zakończenia okresu sprawozdawczego")


def _ident(rows: list[list[str]]) -> dict:
    """Pull provider name + reporting period from a section-1 table. The label
    column varies (some files drop the leading 'Applicability' column), so match
    the (localised) indicator text anywhere in the row; the value is the last
    non-empty cell."""
    out = {"provider": "", "period_start": "", "period_end": ""}
    for r in rows[1:]:
        cells = [c.lower() for c in r]
        non_empty = [i for i, c in enumerate(r) if c]
        if not non_empty:
            continue
        val_idx = non_empty[-1]          # value = last non-empty cell
        val = r[val_idx]
        # everything before the value cell; normalise snake_case labels
        # (e.g. "name_of_the_service_provider") to the spaced form we match on.
        label = " ".join(cells[:val_idx]).replace("_", " ")
        if any(k in label for k in _PROVIDER) and not out["provider"]:
            out["provider"] = val
        elif any(k in label for k in _START) and not out["period_start"]:
            out["period_start"] = val[:10]
        elif any(k in label for k in _END) and not out["period_end"]:
            out["period_end"] = val[:10]
    return out


def _amar_total(rows: list[list[str]]) -> str:
    """AMAR EU total (section 10, scope == TOTAL). Empty for non-VLOP services,
    which leave the value blank (so the last non-empty cell is the 'TOTAL' label
    itself) — hence we only return a genuinely numeric value."""
    for r in rows[1:]:
        if any(c.strip().upper() == "TOTAL" for c in r):
            val = next((c for c in reversed(r) if c), "").replace(",", "").replace(" ", "")
            return val if val.isdigit() else ""
    return ""


def _slugify(name: str) -> str:
    return _re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")


def extract_multi(slug: str, fname: str) -> list[dict]:
    """A single zip holding several products' reports (Miniclip ships one report
    per game, each CSV named '<Game> N_section.csv'). Split by the product prefix
    and emit one extracted dir + manifest entry per product."""
    pairs = read_zipcsv(os.path.join(RAW, fname))  # [(basename, rows), ...]
    products: dict[str, list[tuple[str, list[list[str]]]]] = {}
    for base, rows in pairs:
        m = _re.match(r"(.+?)\s+\d{1,2}[._]", base)
        if m:
            products.setdefault(m.group(1).strip(), []).append((base, rows))
    if not products:  # fail fast if the multi-product naming convention changed
        raise ValueError(f"No '<product> N_section' reports found in {fname}")
    infos = []
    for product, named in sorted(products.items()):
        psl = f"{slug}-{_slugify(product)}"
        infos.append(_write_extracted(psl, fname, "zipmulti", _to_canonical(named)))
    return infos


def _write_extracted(slug: str, fname: str, kind: str,
                     canonical: list[list[list[str]]]) -> dict:
    sheets = [_trim_cols(s) for s in canonical]
    dest = os.path.join(OUT, slug)
    os.makedirs(dest, exist_ok=True)
    section_rows = {}
    for i, name in enumerate(SECTIONS):
        rows = sheets[i] if i < len(sheets) else []
        with open(os.path.join(dest, f"{i + 1:02d}_{name}.csv"), "w",
                  newline="", encoding="utf-8") as f:
            csv.writer(f).writerows(rows)
        section_rows[name] = max(0, len(rows) - 1)
    ident = _ident(sheets[0]) if sheets and sheets[0] else {}
    amar = _amar_total(sheets[9]) if len(sheets) > 9 else ""
    return {
        "platform": slug, "source_file": fname, "format": kind,
        "provider": ident.get("provider", ""),
        "period_start": ident.get("period_start", ""),
        "period_end": ident.get("period_end", ""),
        "amar_eu_total": amar,
        "sections_found": sum(1 for i in range(len(SECTIONS))
                              if i < len(sheets) and sheets[i]),
        "data_rows_by_section": section_rows,
    }


def extract_one(slug: str, fname: str, kind: str) -> dict:
    named = READERS[kind](os.path.join(RAW, fname))
    canonical = (_to_canonical_mapped(named, SHEET_MAP[slug]) if slug in SHEET_MAP
                 else _to_canonical(named))
    return _write_extracted(slug, fname, kind, canonical)


def main() -> None:
    os.makedirs(OUT, exist_ok=True)
    manifest = []
    for slug, (fname, kind) in sorted(SOURCES.items()):
        if not os.path.exists(os.path.join(RAW, fname)):
            print(f"skip {slug}: {fname} not in raw/")
            continue
        infos = extract_multi(slug, fname) if kind == "zipmulti" else [extract_one(slug, fname, kind)]
        manifest.extend(infos)
        for info in infos:
            total = sum(info["data_rows_by_section"].values())
            print(f"{info['platform']:24} {kind:8} {info['sections_found']}/11 sections, "
                  f"{total} data rows")
    with open(os.path.join(HERE, "manifest.json"), "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)
    # Flat cross-platform summary of the headline identification fields.
    with open(os.path.join(HERE, "summary.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["platform", "provider", "period_start", "period_end",
                    "amar_eu_total", "sections_found", "total_data_rows"])
        for m in manifest:
            w.writerow([m["platform"], m["provider"], m["period_start"],
                        m["period_end"], m["amar_eu_total"], m["sections_found"],
                        sum(m["data_rows_by_section"].values())])
    print(f"\nwrote manifest.json + summary.csv ({len(manifest)} platforms) and "
          f"extracted/<platform>/NN_<section>.csv")


if __name__ == "__main__":
    main()
