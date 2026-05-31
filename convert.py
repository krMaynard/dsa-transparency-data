#!/usr/bin/env python3
"""
Convert VLOP DSA report CSVs/xlsx (tables 3-9) to compact JSON for the krMaynard dashboard.
Usage: python3 convert.py
Output: ../krMaynard.github.io/data/vlop-dsa.json
"""

import csv
import json
from pathlib import Path

import openpyxl
import xlrd

REPORTS_DIR = Path(__file__).parent
OUT_FILE = REPORTS_DIR.parent / "krMaynard.github.io" / "data" / "vlop-dsa.json"

# Each entry has name, platform, and either dir (CSV directory) or xlsx (single xlsx file).
SERVICE_DEFS = [
    {"name": "Google Maps",      "dir":  "google/maps",            "platform": "Google", "surfaces": True},
    {"name": "Google Play",      "dir":  "google/play",            "platform": "Google", "surfaces": True},
    {"name": "Google Search",    "dir":  "google/search",          "platform": "Google", "surfaces": True},
    {"name": "Google Shopping",  "dir":  "google/shopping",        "platform": "Google", "surfaces": True},
    {"name": "Multi-Services",   "dir":  "google/multi-services",  "platform": "Google", "surfaces": True},
    {"name": "YouTube",          "dir":  "google/youtube",         "platform": "Google", "surfaces": True},
    {"name": "X",                "dir":  "x",                      "platform": "X"},
    {"name": "TikTok",           "dir":  "tiktok",                 "platform": "TikTok"},
    {"name": "Facebook",         "dir":  "meta/facebook",          "platform": "Meta"},
    {"name": "Instagram",        "dir":  "meta/instagram",         "platform": "Meta"},
    {"name": "Pinterest",        "dir":  "pinterest",              "platform": "Pinterest"},
    {"name": "AliExpress",       "dir":  "aliexpress",             "platform": "Alibaba"},
    {"name": "Amazon",           "dir":  "amazon",                 "platform": "Amazon"},
    {"name": "LinkedIn",         "dir":  "microsoft/linkedin",     "platform": "Microsoft"},
    {"name": "Booking.com",      "dir":  "booking-com",            "platform": "Booking.com"},
    {"name": "App Store",        "xlsx": "apple/app-store.xlsx",   "platform": "Apple"},
    {"name": "Apple Books",      "xlsx": "apple/books.xlsx",       "platform": "Apple"},
    {"name": "iCloud Storage",   "xlsx": "apple/icloud-storage.xlsx", "platform": "Apple"},
    {"name": "Apple Podcasts",   "xlsx": "apple/podcasts.xlsx",    "platform": "Apple"},
    {"name": "Bing",             "xlsx": "microsoft/bing.xlsx",    "platform": "Microsoft"},
    {"name": "SHEIN",            "xlsx": "shein.xlsx",             "platform": "Shein"},
    {"name": "Wikipedia",        "xls":  "wikimedia/wikipedia.xls", "platform": "Wikimedia"},
    {"name": "Wikidata",         "xls":  "wikimedia/wikidata.xls",  "platform": "Wikimedia"},
    {"name": "Wikimedia Commons","xls":  "wikimedia/commons.xls",   "platform": "Wikimedia"},
    {"name": "Wikiversity",      "xls":  "wikimedia/wikiversity.xls", "platform": "Wikimedia"},
    {"name": "Wikivoyage",       "xls":  "wikimedia/wikivoyage.xls", "platform": "Wikimedia"},
    {"name": "Wiktionary",       "xls":  "wikimedia/wiktionary.xls", "platform": "Wikimedia"},
    {"name": "Zalando",          "dir":  "zalando",                "platform": "Zalando"},
    {"name": "Temu",             "dir":  "temu",                   "platform": "Temu"},
    {"name": "Snapchat",         "dir":  "snapchat",               "platform": "Snap"},
    # Adult-content VLOPs (designated Dec 2023 / Jul 2024). Stripchat was de-designated
    # in 2025 and is intentionally excluded.
    {"name": "Pornhub",          "dir":  "PH_DSA_Transparency_Report_FH26_1776436263", "platform": "Aylo"},
    {"name": "XVideos",          "dir":  "XVideos+-+Transparency+report+-+July-December+2025", "platform": "WebGroup Czech Republic"},
    {"name": "XNXX",             "dir":  "XNXX+-+Transparency+report+-+July-December+2025",     "platform": "NKL Associates"},
]

services = []
service_platforms = []
categories = []
sections = []
indicators = []
scopes = []
surfaces = ["All"]  # report surface/breakdown for t6/t7/t8 rows; index 0 = no breakdown

t3_rows = []
t4_rows = []
t5_rows = []
t6_rows = []
t7_rows = []
t8_rows = []
t9_rows = []


def intern(lst, val):
    if val not in lst:
        lst.append(val)
    return lst.index(val)


# Google publishes tables 6/7 as several disjoint sub-reports per service
# (organic "Core", "Ads", and for Search a breakdown by action level). The
# suffix on the filename identifies the surface.
TABLE_STEM = {3: "member_states_orders", 4: "notices", 5: "own_initiative_illegal",
              6: "own_initiative_TC", 7: "appeals_and_recidivism",
              8: "automated_means", 9: "human_resources"}
SURFACE_SUFFIX = {
    "_Ads": "Ads",
    "_Domain_Level_Actions": "Domain-level",
    "_Entity_Level_Actions": "Entity-level",
    "_Host_Level_Actions": "Host-level",
    "_Image_Level_Removals": "Image-level",
    "_Partner_Feed_Domain_Lvl_Action": "Partner feed (domain)",
    "_Partner_Feed_Item_Level_Action": "Partner feed (item)",
    "_Profile_Level_Suspensions": "Profile-level",
    "_Review_Level_Actions": "Review-level",
    "_URL_Level_Removals": "URL-level",
}


def _surface_label(path, n):
    rest = path.stem
    prefix = f"{n}_"
    if rest.startswith(prefix):
        rest = rest[len(prefix):]
    stem = TABLE_STEM.get(n, "")
    if rest == stem:
        return "Core"
    suffix = rest[len(stem):] if rest.startswith(stem) else "_" + rest
    return SURFACE_SUFFIX.get(suffix, suffix.strip("_").replace("_", " ") or "Core")


def table_files(d, n, surfaced=False):
    """Return [(path, surface_label), ...] for table `n`.

    Non-surfaced tables (everything except Google t6/t7) resolve to a single
    canonical file labelled "All"; if several variants exist (e.g. Amazon's
    `_version2`), the shortest filename wins, deterministically. Surfaced
    tables with multiple sub-reports return one entry per surface.
    """
    matches = []
    for pattern in [f"{n}_*.csv", f"*- {n}_*.csv", f"*{n}. *.csv"]:
        matches = sorted(d.glob(pattern))
        if matches:
            break
    if not matches:
        return []
    if not surfaced or n not in (6, 7, 8) or len(matches) == 1:
        return [(min(matches, key=lambda p: len(p.name)), "All")]
    return [(p, _surface_label(p, n)) for p in matches]


def parse_num(val):
    if val is None or val == '':
        return None
    if isinstance(val, (int, float)):
        try:
            return int(val) if float(val) == int(val) else float(val)
        except (ValueError, OverflowError):
            return None
    val = str(val).strip().replace(",", "")
    if val == "":
        return None
    try:
        f = float(val)
        return int(f) if f == int(f) else f
    except ValueError:
        return None


def read_csv(path):
    for enc in ("utf-8-sig", "latin-1"):
        try:
            with open(path, encoding=enc) as f:
                return list(csv.DictReader(f))
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Could not decode {path}")


def read_xls_sheet(xls_path, sheet_name):
    """Read a named sheet from a legacy .xls file as a list of dicts."""
    wb = xlrd.open_workbook(str(xls_path))
    if sheet_name not in wb.sheet_names():
        return []
    ws = wb.sheet_by_name(sheet_name)
    if ws.nrows < 1:
        return []
    headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
    result = []
    for r in range(1, ws.nrows):
        d = {}
        for c, h in enumerate(headers):
            v = ws.cell_value(r, c)
            d[h] = v if v != '' else None
        result.append(d)
    return result


def read_xlsx_sheet(xlsx_path, sheet_name):
    """Read a named sheet from an xlsx as a list of dicts (same contract as read_csv)."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    result = []
    for row in rows[1:]:
        d = {}
        for i, h in enumerate(headers):
            d[h] = row[i] if i < len(row) else None
        result.append(d)
    return result


def get(row, *keys):
    for k in keys:
        if k in row:
            return row[k]
        k2 = k.strip()
        for rk in row:
            if rk.strip() == k2:
                return row[rk]
    return None


def any_numeric(vals):
    return any(v is not None and v != 0 for v in vals)


def process_t3(svc_idx, rows):
    for row in rows:
        cat = str(get(row, "Category of illegal content") or "").strip()
        if not cat:
            continue
        scope_val = str(get(row, "Scope") or "").strip()
        orders_act = parse_num(get(row, "Number of orders to act against illegal content received"))
        items = parse_num(get(row, "Number of specific items of information included in the total number of orders to act against illegal content"))
        orders_info = parse_num(get(row, "Number of orders to provide information"))
        nums = [orders_act, items, orders_info]
        if not any_numeric(nums):
            continue
        t3_rows.append([svc_idx, intern(categories, cat), intern(scopes, scope_val),
                         orders_act, items, orders_info])


def process_t4(svc_idx, rows):
    for row in rows:
        cat = str(get(row, "Category of illegal content") or "").strip()
        if not cat:
            continue
        notices = parse_num(get(row, "Number of notices received ", "Number of notices received"))
        tf_notices = parse_num(get(row, "Number of notices received from Trusted flaggers"))
        items = parse_num(get(row, "Number of specific items of information included in the total number of notices"))
        tf_items = parse_num(get(row, "Number of specific items of information included in the total number of notices by Trusted Flaggers (Trusted Flagger notices)"))
        median = parse_num(get(row, "Median time to take action"))
        tf_median = parse_num(get(row, "Median time to take action (Trusted Flagger notices)"))
        act_law = parse_num(get(row, "Number of actions taken on the basis of the law"))
        tf_act_law = parse_num(get(row, "Number of actions taken on the basis of the law (Trusted Flagger notices)"))
        act_tc = parse_num(get(row, "Number of actions taken on the basis of the terms and conditions of the service"))
        tf_act_tc = parse_num(get(row, "Number of actions taken on the basis of the terms and conditions of the service (Trusted Flagger notices)"))
        nums = [notices, tf_notices, items, act_law, act_tc]
        if not any_numeric(nums):
            continue
        t4_rows.append([svc_idx, intern(categories, cat),
                         notices, tf_notices, items, tf_items,
                         median, tf_median,
                         act_law, tf_act_law, act_tc, tf_act_tc])


def process_t5_t6(svc_idx, rows, cat_key, out_rows, surface=None):
    for row in rows:
        cat = str(get(row, cat_key) or "").strip()
        if not cat:
            continue
        measures = parse_num(get(row, "Number of measures taken at the provider's own initiative ", "Number of measures taken at the provider's own initiative"))
        automated = parse_num(get(row, "Number of measures taken after detection with solely automated means ", "Number of measures taken after detection with solely automated means"))
        removal = parse_num(get(row, "Visibility restriction Removal"))
        disable = parse_num(get(row, "Visibility restriction Disable"))
        demoted = parse_num(get(row, "Visibility restriction Demoted"))
        age_restr = parse_num(get(row, "Visibility restriction Age restricted"))
        interaction = parse_num(get(row, "Visibility restriction Interaction restricted"))
        labelled = parse_num(get(row, "Visibility restriction Labelled ", "Visibility restriction Labelled"))
        vis_other = parse_num(get(row, "Visibility restriction Other"))
        mon_susp = parse_num(get(row, "Monetary restriction Suspension"))
        mon_term = parse_num(get(row, "Monetary restriction Termination"))
        mon_other = parse_num(get(row, "Monetary restriction Other"))
        svc_susp = parse_num(get(row, "Provision of the service Suspension"))
        svc_term = parse_num(get(row, "Provision of the service Termination"))
        acc_susp = parse_num(get(row, "Account restriction Suspension"))
        acc_term = parse_num(get(row, "Account restriction Termination"))
        nums = [measures, automated, removal, disable, demoted, age_restr, interaction,
                labelled, vis_other, mon_susp, mon_term, mon_other, svc_susp, svc_term,
                acc_susp, acc_term]
        if not any_numeric(nums):
            continue
        out = [svc_idx, intern(categories, cat),
               measures, automated,
               removal, disable, demoted, age_restr, interaction, labelled, vis_other,
               mon_susp, mon_term, mon_other,
               svc_susp, svc_term,
               acc_susp, acc_term]
        if surface is not None:
            out.append(intern(surfaces, surface))
        out_rows.append(out)


def process_t7(svc_idx, rows, surface="All"):
    for row in rows:
        section = str(get(row, "Section") or "").strip()
        indicator = str(get(row, "Indicator") or "").strip()
        scope_val = str(get(row, "Scope") or "").strip()
        value = parse_num(get(row, "Value"))
        if not section or not indicator or value is None:
            continue
        t7_rows.append([svc_idx, intern(sections, section),
                         intern(indicators, indicator),
                         intern(scopes, scope_val), value,
                         intern(surfaces, surface)])


def process_t8(svc_idx, rows, surface="All"):
    for row in rows:
        section = str(get(row, "Section") or "").strip()
        indicator = str(get(row, "Indicator") or "").strip()
        scope_val = str(get(row, "Scope") or "").strip()
        raw_val = get(row, "Value")
        value = parse_num(raw_val)
        if value is None and isinstance(raw_val, str) and " - " in raw_val:
            parts = raw_val.split(" - ", 1)
            v1, v2 = parse_num(parts[0].strip()), parse_num(parts[1].strip())
            if v1 is not None and v2 is not None:
                value = (v1 + v2) / 2
        if not section or not indicator or value is None:
            continue
        t8_rows.append([svc_idx, intern(sections, section),
                         intern(indicators, indicator),
                         intern(scopes, scope_val), value,
                         intern(surfaces, surface)])


def process_t9(svc_idx, rows):
    last_section = ""
    last_indicator = ""
    for row in rows:
        section = str(get(row, "Section") or "").strip()
        indicator = str(get(row, "Indicator") or "").strip()
        if section:
            last_section = section
        else:
            section = last_section
        if indicator:
            last_indicator = indicator
        else:
            indicator = last_indicator
        scope_val = str(get(row, "Scope") or "").strip()
        value = parse_num(get(row, "Value"))
        if not section or not indicator or value is None:
            continue
        t9_rows.append([svc_idx, intern(sections, section),
                         intern(indicators, indicator),
                         intern(scopes, scope_val), value])


def build_category_labels():
    google_maps_dir = next((s["dir"] for s in SERVICE_DEFS if "Google Maps" in s.get("name", "")), None)
    if not google_maps_dir:
        return {}
    label_path = REPORTS_DIR / google_maps_dir / "2_categories_names.csv"
    labels = {}
    if label_path.exists():
        for row in read_csv(label_path):
            code = (row.get("Category of illegal content / incompatible with the terms and conditions") or "").strip()
            desc = (row.get("Category description") or "").strip()
            if code and desc:
                labels[code] = desc
    return labels


def process_service_from_dir(svc_idx, d):
    svc_name = services[svc_idx]
    svc_def = next((s for s in SERVICE_DEFS if s["name"] == svc_name), {})
    surfaced = svc_def.get("surfaces", False)
    for path, _ in table_files(d, 3):
        process_t3(svc_idx, read_csv(path))

    for path, _ in table_files(d, 4):
        process_t4(svc_idx, read_csv(path))

    for path, _ in table_files(d, 5):
        process_t5_t6(svc_idx, read_csv(path),
                       "Category of illegal content", t5_rows)

    for path, surface in table_files(d, 6, surfaced):
        process_t5_t6(svc_idx, read_csv(path),
                       "Category of incompatibility with the provider's terms and conditions",
                       t6_rows, surface=surface)

    for path, surface in table_files(d, 7, surfaced):
        process_t7(svc_idx, read_csv(path), surface=surface)

    for path, surface in table_files(d, 8, surfaced):
        process_t8(svc_idx, read_csv(path), surface=surface)

    for path, _ in table_files(d, 9):
        process_t9(svc_idx, read_csv(path))


def process_service_from_xls(svc_idx, xls_path):
    process_t3(svc_idx, read_xls_sheet(xls_path, "3_member_states_orders"))
    process_t4(svc_idx, read_xls_sheet(xls_path, "4_notices"))
    process_t5_t6(svc_idx, read_xls_sheet(xls_path, "5_own_initiative_illegal"),
                  "Category of illegal content", t5_rows)
    process_t5_t6(svc_idx, read_xls_sheet(xls_path, "6_own_initiative_TC"),
                  "Category of incompatibility with the provider's terms and conditions",
                  t6_rows, surface="All")
    process_t7(svc_idx, read_xls_sheet(xls_path, "7_appeals_and_recidivism"))
    process_t8(svc_idx, read_xls_sheet(xls_path, "8_automated_means"))
    process_t9(svc_idx, read_xls_sheet(xls_path, "9_human_resources"))


def process_service_from_xlsx(svc_idx, xlsx_path):
    process_t3(svc_idx, read_xlsx_sheet(xlsx_path, "3_member_states_orders"))
    process_t4(svc_idx, read_xlsx_sheet(xlsx_path, "4_notices"))
    process_t5_t6(svc_idx, read_xlsx_sheet(xlsx_path, "5_own_initiative_illegal"),
                  "Category of illegal content", t5_rows)
    process_t5_t6(svc_idx, read_xlsx_sheet(xlsx_path, "6_own_initiative_TC"),
                  "Category of incompatibility with the provider's terms and conditions",
                  t6_rows, surface="All")
    process_t7(svc_idx, read_xlsx_sheet(xlsx_path, "7_appeals_and_recidivism"))
    process_t8(svc_idx, read_xlsx_sheet(xlsx_path, "8_automated_means"))
    process_t9(svc_idx, read_xlsx_sheet(xlsx_path, "9_human_resources"))


def main():
    for svc_def in SERVICE_DEFS:
        svc_name = svc_def["name"]
        platform = svc_def["platform"]
        print(f"Processing {svc_name}...")

        if "dir" in svc_def:
            d = REPORTS_DIR / svc_def["dir"]
            if not d.exists():
                print(f"  WARNING: missing {d}")
                continue
            svc_idx = intern(services, svc_name)
            if len(service_platforms) <= svc_idx:
                service_platforms.append(platform)
            process_service_from_dir(svc_idx, d)

        elif "xlsx" in svc_def:
            xlsx_path = REPORTS_DIR / svc_def["xlsx"]
            if not xlsx_path.exists():
                print(f"  WARNING: missing {xlsx_path}")
                continue
            svc_idx = intern(services, svc_name)
            if len(service_platforms) <= svc_idx:
                service_platforms.append(platform)
            process_service_from_xlsx(svc_idx, xlsx_path)

        elif "xls" in svc_def:
            xls_path = REPORTS_DIR / svc_def["xls"]
            if not xls_path.exists():
                print(f"  WARNING: missing {xls_path}")
                continue
            svc_idx = intern(services, svc_name)
            if len(service_platforms) <= svc_idx:
                service_platforms.append(platform)
            process_service_from_xls(svc_idx, xls_path)

    category_labels = build_category_labels()

    out = {
        "meta": {
            "period": "2025-07-01/2025-12-31",
            "generated": "2026-05-13",
        },
        "services": services,
        "service_platforms": service_platforms,
        "categories": categories,
        "category_labels": category_labels,
        "sections": sections,
        "indicators": indicators,
        "scopes": scopes,
        "surfaces": surfaces,
        "t3": t3_rows,
        "t4": t4_rows,
        "t5": t5_rows,
        "t6": t6_rows,
        "t7": t7_rows,
        "t8": t8_rows,
        "t9": t9_rows,
    }

    OUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_FILE, "w") as f:
        json.dump(out, f, separators=(",", ":"))

    print(f"\nWritten to {OUT_FILE}")
    print(f"  services: {services}")
    print(f"  categories: {len(categories)}")
    print(f"  t3 rows: {len(t3_rows)}")
    print(f"  t4 rows: {len(t4_rows)}")
    print(f"  t5 rows: {len(t5_rows)}")
    print(f"  t6 rows: {len(t6_rows)}")
    print(f"  t7 rows: {len(t7_rows)}")
    print(f"  t8 rows: {len(t8_rows)}")
    print(f"  t9 rows: {len(t9_rows)}")


if __name__ == "__main__":
    main()
